import openpyxl
import os
import sys


def getWorkBook(filename):
    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()
    return wb


def saveWorkBook(wb, filename):
    wb.save(filename)


def getColumnData(sheet, column):
    data = []
    for row in range(2,sheet.max_row+1):
        data.append(getCellValue(row, column, sheet))
    return data


def getRowData(sheet, row):
    data = []
    for column in range(2,sheet.max_column+1):
        data.append(getCellValue(row, column, sheet))
    return data


def getSheetByName(wb, sheetname):
    if sheetExists(wb, sheetname):
        return wb[sheetname]
    print("{} does not exist".format(sheetname))
    sys.exit(0)


def sheetExists(wb, sheetname):
    for name in wb.sheetnames:
        if name == sheetname:
            return True
    return False


def removeSheets(wb):
    sheets = wb.sheetnames
    for name in sheets:
        if (name != "Point Distribution 2022"):
            sheet = wb[name]
            wb.remove(sheet)
    return wb


def writeToCell(row, col, sheet, value):
    sheet.cell(row=row, column=col).value = value
    return sheet


def getCellValue(row, col, sheet):
    return sheet.cell(row=row, column=col).value


def createSheet(wb, sheetName):
    return wb.create_sheet(sheetName)


def createTeamsRankedSheet(wb):
    sheet = createSheet(wb, "Teams Ranked")
    sheet = writeToCell(1, 1, sheet, 'Team Name')
    sheet = writeToCell(1, 2, sheet, 'Player One')
    sheet = writeToCell(1, 3, sheet, 'Player Two')
    sheet = writeToCell(1, 4, sheet, 'Points')
    return wb


def createTeamsSheet(wb):
    sheet = createSheet(wb, "Teams")
    sheet = writeToCell(1, 1, sheet, 'Team Name')
    sheet = writeToCell(1, 2, sheet, 'Player One')
    sheet = writeToCell(1, 3, sheet, 'Player Two')
    sheet = writeToCell(1, 4, sheet, 'Total Points')
    return wb


def createPlayersRankedSheet(wb):
    sheet = createSheet(wb, "Players Ranked")
    sheet = writeToCell(1, 1, sheet, 'Players')
    sheet = writeToCell(1, 2, sheet, 'Points')
    return wb


def createPlayersSheet(wb):
    sheet = createSheet(wb, "Players")
    sheet = wb['Players']
    sheet = writeToCell(1, 1, sheet, 'Player')
    sheet = writeToCell(1, 2, sheet, 'Total')
    sheet = writeToCell(1, 3, sheet, 'Result 1')
    sheet = writeToCell(1, 4, sheet, 'Result 2')
    sheet = writeToCell(1, 5, sheet, 'Result 3')
    return wb
