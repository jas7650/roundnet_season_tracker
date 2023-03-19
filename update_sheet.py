import os
import sys
import openpyxl
import scrape_utils

RANK = 1
TEAM = 2
PLAYER_ONE = 3
PLAYER_TWO = 4
POINTS = 5

CONTENDER = 0
CHALLENGER = 1 
MAJOR = 2
CHAMPIONSHIP_PRO = 3
CHAMPIONSHIP_PREMIER = 4


def main():

    if os.path.exists('roundnet_tracking_2023.xlsx'):
        wb = openpyxl.load_workbook('roundnet_tracking_2023.xlsx')
        wb = removeSheets(wb)
    else:
        wb = openpyxl.Workbook()
        wb = removeSheets(wb)
   
    tourney_path = os.path.join(os.getcwd(), "tournaments")
    
    wb = createPlayersSheet(wb)
    wb = createPlayersRankedSheet(wb)
    wb = createTeamsSheet(wb)
    wb = createTeamsRankedSheet(wb)
    wb = readYearDirectory(wb, tourney_path, 2022)  
    wb = readYearDirectory(wb, tourney_path, 2023)
    wb = updatePlayerTotals(wb)
    wb = writePlayersRankedSheet(wb)
    wb = writeTeamsSheet(wb)
    wb = writeTeamsRankedSheet(wb)
           
    # path = os.path.join(os.getcwd(), "tournaments")
    # path = os.path.join(path, "2023")

    wb.save('roundnet_tracking_2023.xlsx')
    # scrapeWeb()

# Players - Name, Points
# Teams - Name, Player Names, Points


def readYearDirectory(wb, tourney_path, year):
    print("Year: {}".format(year))
    year_path = os.path.join(tourney_path, str(year))
    if (os.path.exists(year_path)):
        wb = readDirectory(wb, os.path.join(year_path, os.path.join("championship", "pro")), CHAMPIONSHIP_PRO)
        wb = readDirectory(wb, os.path.join(year_path, "major"), MAJOR)
        wb = readDirectory(wb, os.path.join(year_path, os.path.join("championship", "premier")), CHAMPIONSHIP_PREMIER)
        wb = readDirectory(wb, os.path.join(year_path, "challenger"), CHALLENGER)
        wb = readDirectory(wb, os.path.join(year_path, "contender"), CONTENDER)
    return wb


def readDirectory(wb, path, TOURNAMENT_TYPE):
    if (os.path.exists(path)):
        for filename in os.listdir(path):
            file = os.path.join(path, filename)
            if os.path.isfile(file):
                ranks, teams = scrape_utils.scrapeFile(file)
                if (TOURNAMENT_TYPE == CHAMPIONSHIP_PREMIER):
                    for i in range(len(ranks)):
                        ranks[i] = ranks[i]+16
                location = os.path.splitext(os.path.basename(file))[0]
                location = location.replace("_", " ")
                print("{}".format(location))
                ideal_points = getPointsArray(2022, TOURNAMENT_TYPE)
                old_points = getActualPoints(ranks, ideal_points)
                points = []
                for value in old_points:
                    points.append(value/2.0)
                writeTournamentSheet(wb, ranks, teams, points, location+" 2022")
                writePlayerSheet(wb, teams, points, location)
    return wb


def updatePlayerTotals(wb):
    sheet = wb['Players']
    players = getPlayers(sheet)
    locations = getExistingLocations(sheet)
    for i in range(len(players)):
        points = [0, 0, 0]
        row = i + 2
        for j in range(len(locations)):
            col = j + 6
            value = sheet.cell(row=row, column=col).value
            if value is None:
                value = 0
            if value > points[0]:
                points[2] = points[1]
                points[1] = points[0]
                points[0] = value
            elif value > points[1]:
                points[2] = points[1]
                points[1] = value
            elif value > points[2]:
                points[2] = points[1]
        sheet.cell(row=row, column=2).value = points[0] + points[1] + points[2]
        sheet.cell(row=row, column=3).value = points[0]
        sheet.cell(row=row, column=4).value = points[1]
        sheet.cell(row=row, column=5).value = points[2]
    return wb


def contains(list, value):
    i = 0
    while (i < len(list)):
        if (list[i] == value):
            return i
        i += 1
    return -1


def removeSheets(wb):
    sheets = wb.sheetnames
    for name in sheets:
        if (name != "Point Distribution 2022"):
            sheet = wb[name]
            wb.remove(sheet)
    return wb


def sortPlayers(players, points):
    players_copy = players
    points_copy = points
    players_sorted = []
    points_sorted = []
    while len(players_copy) > 0:
        largest_points = 0
        largest_player = ""
        index = 0
        for i in range(len(points_copy)):
            amount = points_copy[i]
            if amount > largest_points:
                index = i
                largest_points = amount
                largest_player = players_copy[i]
        del players_copy[index]
        del points_copy[index]
        players_sorted.append(largest_player)
        points_sorted.append(largest_points)
    return players_sorted, points_sorted

def sortTeams(teams, players, points):
    teams_copy = teams
    teams_sorted = []
    points_sorted = []

    while len(teams_copy) > 0:
        largest_points = 0
        largest_team = teams_copy[0]
        index = 0
        for i in range(len(teams_copy)):
            team = teams_copy[i]
            p1_points = getPlayerPoints(team[1], players, points)
            p2_points = getPlayerPoints(team[2], players, points)
            amount = p1_points + p2_points
            if amount > largest_points:
                index = i
                largest_points = amount
                largest_team = teams_copy[i]
        del teams_copy[index]
        teams_sorted.append(largest_team)
        points_sorted.append(largest_points)
        
    return teams_sorted, points_sorted


def getPlayerPoints(player, players, points):
    for i in range(len(players)):
        if (players[i] == player):
            return points[i]
    return 0


def getPointsArray(year, TOURNAMENT_TYPE):
    if os.path.exists('point_distribution.xlsx'):
        wb = openpyxl.load_workbook('point_distribution.xlsx')
    else:
        print("Point distrubution excel file must exist")
        sys.exit(0)
    points = []
    if (year == 2022):
        sheet = wb['Point Distribution 2022']
    else:
        sheet = wb['Point Distribution 2023']
    if (TOURNAMENT_TYPE == CHAMPIONSHIP_PREMIER):
        column = CHAMPIONSHIP_PRO + 2 
    else:
        column = TOURNAMENT_TYPE + 2
    for row in range(2,sheet.max_row+1):
        points.append(sheet.cell(row=row, column=column).value)
    return points


def getExistingLocations(sheet):
    locations = []
    row = 1

    for col in range(6,sheet.max_column+1):
        if (sheet.cell(row=row, column=col).value == "None"):
            break
        locations.append(sheet.cell(row=row, column=col).value)
    return locations


def getLocation(file):
    index = str(file).find("_")
    text = str(file)[:index]
    if (text == "champ"):
        text = str(file)[index+1:]
        index = text.find("_")
        text = text[index+1:]
        text = os.path.splitext(text)[0]
        return text
    index = str(file).find("_")
    text = str(file)[index+1:]
    text = os.path.splitext(text)[0]
    return text


def getColumn(sheet, location):
    for col in range(6,sheet.max_column+1):
        if (sheet.cell(row=1, column=col).value == location):
            return col
    return sheet.max_column+1


def getActualPoints(ranks, points):
    previousRank = 0
    average = 0
    i = 0
    actual_points = []
    if (len(ranks) != len(points)):
        points = points[ranks[0]-1:]
    while (i < len(ranks)):
        rank = ranks[i]
        if (rank != previousRank):
            average = getAverageValue(ranks, points, i)

        actual_points.append(average)
        previousRank = rank
        i = i + 1
    return actual_points


def getAverageValue(ranks, points, index):
    if (index == len(ranks)-1):
        return ranks[index]
    i = index+1
    numValues = 1
    sum = points[index]

    while (ranks[i] == ranks[index]):
        sum += points[i]
        numValues += 1
        i += 1
        if (i == len(ranks)):
            break
    return sum/numValues


def getTeamsFromTournament(sheet):
    team_names = []
    player_ones = []
    player_twos = []
    teams = []
    for row in range(2, sheet.max_row+1):
        for column in "BCD":
            cell_name = "{}{}".format(column, row)
            if (column == "B"):
                team_names.append(sheet[cell_name].value)
            if (column == "C"):
                player_ones.append(sheet[cell_name].value)
            if (column == "D"):
                player_twos.append(sheet[cell_name].value)
    for i in range(len(team_names)):
        team = [team_names[i], player_ones[i], player_twos[i]]
        teams.append(team)
    return teams


def getAllTeams(wb):
    team_names = []
    player_ones = []
    player_twos = []
    teams = []
    sheet = wb['Teams']
    for row in range(2, sheet.max_row+1):
        for column in "ABC":
            cell_name = "{}{}".format(column, row)
            if (column == "A"):
                team_names.append(sheet[cell_name].value)
            if (column == "B"):
                player_ones.append(sheet[cell_name].value)
            if (column == "C"):
                player_twos.append(sheet[cell_name].value)
    for i in range(len(team_names)):
        team = [team_names[i], player_ones[i], player_twos[i]]
        teams.append(team)
    return teams


def getPlayers(sheet):
    players = []
    for row in range(2,sheet.max_row+1):
        for column in "A":
            cell_name = "{}{}".format(column, row)
            players.append(sheet[cell_name].value)
    return players


def getPoints(sheet):
    points = []
    for row in range(2,sheet.max_row+1):
        for column in "B":
            cell_name = "{}{}".format(column, row)
            points.append(sheet[cell_name].value)
    return points


def writePlayerSheet(wb, teams, points, location):
    sheet = wb['Players']
    column = getColumn(sheet, location)

    sheet.cell(row = 1, column = column).value = location

    for playernum in range(2):
        i = 1
        while (i <= len(teams[playernum+1])):
            players = getPlayers(sheet)
            player_row = contains(players, teams[playernum+1][i-1])
            if (player_row != -1):
                row = player_row+2
            else:
                row = len(players)+2
            sheet.cell(row = row, column = column).value = points[i-1]
            sheet.cell(row = row, column = 1).value = teams[playernum+1][i-1]
            i += 1
    

def writeTeamsSheet(wb):
    sheets = wb.sheetnames
    teams_sheet = wb['Teams']
    for name in sheets:
        if (str(name).find("2022") != -1 and name != "Point Distribution 2022"):
            sheet = wb[name]
            teams = getTeamsFromTournament(sheet)
            all_teams = getAllTeams(wb)
            i = len(all_teams)
            for team in teams:
                skip = False
                all_teams = getAllTeams(wb)
                for team_check in all_teams:
                    if (team_check[1] == team[1] and team_check[2] == team[2]):
                        skip = True
                        break    
                    if (team_check[1] == team[2] and team_check[2] == team[1]):
                        skip = True
                        break
                if (skip == True):
                    continue
                teams_sheet.cell(row=i+2,column=1).value = team[0]
                teams_sheet.cell(row=i+2,column=2).value = team[1]
                teams_sheet.cell(row=i+2,column=3).value = team[2]
                i += 1
    return wb


def writeTournamentSheet(wb, ranks, teams, points, filename):
    if not (filename in wb.sheetnames):
        sheet = wb.create_sheet(title=filename)
        offset = 0
    else:
        sheet = wb[filename]
        offset = len(getPlayers(sheet))

    i = 1
    sheet.cell(row=1, column=RANK).value = 'Rank'
    sheet.cell(row=1, column=TEAM).value = 'Team'
    sheet.cell(row=1, column=PLAYER_ONE).value = 'Player One'
    sheet.cell(row=1, column=PLAYER_TWO).value = 'Player Two'
    sheet.cell(row=1, column=POINTS).value = 'Points Awarded'

    while (i <= len(ranks)):
        sheet.cell(row=i+offset+1, column=RANK).value = ranks[i-1]
        sheet.cell(row=i+offset+1, column=TEAM).value = teams[0][i-1]
        sheet.cell(row=i+offset+1, column=PLAYER_ONE).value = teams[1][i-1]
        sheet.cell(row=i+offset+1, column=PLAYER_TWO).value = teams[2][i-1]
        sheet.cell(row=i+offset+1, column=POINTS).value = points[i-1]
        i = i + 1


def writePlayersRankedSheet(wb):
    players_sheet = wb["Players"]
    sheet = wb["Players Ranked"]
    players = getPlayers(players_sheet)
    points = getPoints(players_sheet)
    players, points = sortPlayers(players, points)
    numPlayers = len(players)
    for i in range(numPlayers):
        sheet.cell(row=i+2,column=1).value = players[i]
        sheet.cell(row=i+2,column=2).value = points[i]
    return wb


def writeTeamsRankedSheet(wb):
    sheet = wb["Teams Ranked"]
    players_sheet = wb["Players"]
    teams = getAllTeams(wb)
    players = getPlayers(players_sheet)
    points = getPoints(players_sheet)

    teams, points = sortTeams(teams, players, points)
    numTeams = len(teams)
    for i in range(numTeams):
        sheet.cell(row=i+2,column=1).value = teams[i][0]
        sheet.cell(row=i+2,column=2).value = teams[i][1]
        sheet.cell(row=i+2,column=3).value = teams[i][2]
        sheet.cell(row=i+2,column=4).value = points[i]
    return wb


def createTeamsRankedSheet(wb):
    sheet = wb.create_sheet("Teams Ranked")
    sheet.cell(row=1, column=1).value = "Team Name"
    sheet.cell(row=1, column=2).value = "Player One"
    sheet.cell(row=1, column=3).value = "Player Two"
    sheet.cell(row=1, column=4).value = "Points"
    return wb


def createTeamsSheet(wb):
    sheet = wb.create_sheet("Teams")
    sheet.cell(row=1, column=1).value = "Team Name"
    sheet.cell(row=1, column=2).value = "Player One"
    sheet.cell(row=1, column=3).value = "Player Two"
    sheet.cell(row=1, column=4).value = "Total Points"
    return wb


def createPlayersRankedSheet(wb):
    sheet = wb.create_sheet("Players Ranked")
    sheet.cell(row=1, column=1).value = "Players"
    sheet.cell(row=1, column=2).value = "Points"
    return wb


def createPlayersSheet(wb):
    wb.create_sheet("Players")
    sheet = wb['Players']
    sheet.cell(row=1,column=1).value = "Player"
    sheet.cell(row=1,column=2).value= "Total"
    sheet.cell(row=1,column=3).value= "Result 1"
    sheet.cell(row=1,column=4).value= "Result 2"
    sheet.cell(row=1,column=5).value= "Result 3"
    return wb


if __name__ == "__main__":
    main()
